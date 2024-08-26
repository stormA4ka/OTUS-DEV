import os
import msvcrt
import sys
from openpyxl import load_workbook, Workbook
from datetime import datetime
import pandas as pd


def is_file_exists(dir_path: str, file_name: str) -> None:
    """
    Проверяем существует ли файл excel с именем file_name.
    Если нет, то создаем его с заголовками ["ID", "ФИО", "Телефон", "Комментарий", "Дата добавления"]
    Args:
        dir_path:
        file_name:

    Returns:

    """

    # Формируем полный путь к файлу
    file_path = os.path.join(dir_path, file_name)

    # Проверяем, существует ли файл
    if not os.path.exists(file_path):
        # Если директории нет, создаем её
        if not os.path.exists(dir_path):
            os.makedirs(dir_path)

        # Создаем новый Excel файл
        wb = Workbook()
        ws = wb.active

        # Добавляем заголовки
        cols = ["ID", "ФИО", "Телефон", "Комментарий", "Дата добавления"]
        ws.append(cols)

        # Сохраняем файл
        wb.save(file_path)
        print(f"Файл {file_path} сейчас был создан и содержит пустую таблицу с заголовками {cols}")
    else:
        print(f"Файл {file_path} уже существует.")


def is_file_locked(file_path: str) -> None:
    """
    Проверяем открыт ли файл с именем file_name в другой программе
    Args:
        file_path:

    Returns:

    """
    try:
        with open(file_path, 'r+b') as f:
            msvcrt.locking(f.fileno(), msvcrt.LK_RLCK, 1)
            return False
    except IOError:
        return True


def insert_row(dir_path: str, file_name: str) -> None:
    """
    Добавляем запись (абонента) в файл file_name.

    Вводим поля вручную ["ФИО", "Телефон", "Комментарий"]
    Заполняются автоматически ["ID", "Дата добавления"]
    Args:
        dir_path:
        file_name:

    Returns:

    """

    # Формируем полный путь к файлу
    file_path = os.path.join(dir_path, file_name)

    # Проверяем есть ли такой файл, если нет, то создаем его
    is_file_exists(dir_path, file_name)

    # Проверяем, заблокирован ли файл
    if is_file_locked(file_path):
        print(
            f"Ошибка: Файл {file_path} заблокирован (открыт в другой программе). Пожалуйста, закройте файл и повторите попытку.")
        return

    # Запрашиваем данные у пользователя
    fio = input("Введите ФИО: ")
    phone = input("Введите Телефон: ")
    comment = input("Введите Комментарий: ")

    try:
        # Загружаем существующий Excel файл
        wb = load_workbook(file_path)
        ws = wb.active

        # Ищем максимальный ID
        max_id = -1
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            if row[0] is not None:
                max_id = max(max_id, row[0])

        # Присваиваем новый ID
        new_id = max_id + 1

        # Получаем текущую дату и время
        current_datetime = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        # Вставляем новые данные
        ws.append([new_id, fio, phone, comment, current_datetime])

        # Сохраняем файл
        wb.save(file_path)
        print(f"Данные добавлены с ID {new_id}.")
    except Exception as e:
        print(f"Произошла ошибка: {e}")


def info_book(dir_path: str, file_name: str) -> None:
    """
    Выводим информацию об указанной базе.
    Args:
        dir_path:
        file_name:

    Returns:

    """

    # Формируем полный путь к файлу
    file_path = os.path.join(dir_path, file_name)
    df = pd.read_excel(file_path)
    print(df.info())


def read_book(dir_path: str, file_name: str) -> None:
    """
    Читаем базу.
    Args:
        dir_path:
        file_name:

    Returns:

    """
    # Формируем полный путь к файлу
    pd.set_option('display.max_rows', None)
    pd.set_option('display.max_columns', None)
    pd.set_option('display.width', None)

    file_path = os.path.join(dir_path, file_name)
    df = pd.read_excel(file_path)
    print(df)


def search_rows(dir_path: str, file_name: str) -> None:
    """
    Выбор колонки для поиска и поиск строк по введенным значениям.

    :param df: DataFrame для поиска
    :return: Список ID найденных строк
    """

    file_path = os.path.join(dir_path, file_name)
    df = pd.read_excel(file_path)
    cols = ["ID", "ФИО", "Телефон", "Комментарий", "Дата добавления"]
    print("Доступные колонки для поиска:")
    for i, col in enumerate(cols):
        print(f"{i + 1}. {col}")

    # Запрашиваем выбор колонки
    col_choice = int(input("Выберите номер колонки для поиска: ")) - 1
    selected_col = cols[col_choice]

    # Запрашиваем ввод значений для выбранной колонки
    values_input = input(f"Введите значения для поиска в колонке '{selected_col}' (через запятую, если несколько): ")
    values = [value.strip() for value in values_input.split(",") if value.strip()]

    # Преобразуем значения в соответствии с типом данных в колонке
    if selected_col == "ID":
        try:
            values = [int(value) for value in values]
        except ValueError:
            print("Ошибка: значения в колонке 'ID' должны быть числами.")
            return []

    # Создаем условие для фильтрации
    condition = df[selected_col].isin(values)

    # Находим строки, соответствующие условию
    rows_to_remove = df[condition]

    # Возвращаем список ID найденных строк
    return rows_to_remove, rows_to_remove['ID'].tolist()


def remove_rows(dir_path: str, file_name: str) -> None:
    """
    Находит и удаляет строки из DataFrame на основе значений в указанных колонках.
    Поиск осуществляется по выбранному пользователем полю.

    :param dir_path: Путь к директории с файлом
    :param file_name: Имя файла
    :return: DataFrame без указанных строк
    """

    file_path = os.path.join(dir_path, file_name)
    df = pd.read_excel(file_path)

    # Проверяем, заблокирован ли файл
    if is_file_locked(file_path):
        print(
            f"Ошибка: Файл {file_path} заблокирован (открыт в другой программе). Пожалуйста, закройте файл и повторите попытку.")
        return

    # Вызываем функцию поиска строк
    rows_to_edit, ids_to_remove = search_rows(dir_path, file_name)

    # Проверяем, есть ли строки для удаления
    if ids_to_remove:
        # Находим строки, соответствующие найденным ID
        rows_to_remove = df[df['ID'].isin(ids_to_remove)]
        print("Найдены следующие записи:")
        print(rows_to_remove)
        confirm = input("Точно удалить эти записи? (y/n): ")
        if confirm.lower() == 'y':
            # Удаляем строки, соответствующие найденным ID
            filtered_df = df[~df['ID'].isin(ids_to_remove)]
            print(f"Удалено строк: {len(ids_to_remove)}")
            filtered_df.to_excel(file_path, sheet_name='Sheet', index=False)
            return filtered_df
        else:
            print("Записи не удалены.")
            return df
    else:
        print("Записи не найдены.")
        return df


def edit_row(dir_path: str, file_name: str) -> None:
    """
    Находит и редактирует строку в DataFrame на основе значения в указанной колонке.
    После редактирования сохраняет изменения в файл Excel.

    :param dir_path: Путь к директории с файлом
    :param file_name: Имя файла
    :return: DataFrame с отредактированной строкой
    """
    file_path = os.path.join(dir_path, file_name)
    df = pd.read_excel(file_path)

    # Вызываем функцию поиска строки для редактирования
    rows_to_edit, ids_to_edit = search_rows(dir_path, file_name)

    # Проверяем, есть ли строки для редактирования
    if not rows_to_edit.empty:
        print("Найдены следующие записи:")
        print(rows_to_edit)
        confirm = input("Точно редактировать эти записи? (y/n): ")
        if confirm.lower() == 'y':
            # Запрашиваем новые значения для каждой колонки
            editable_cols = ["ФИО", "Телефон", "Комментарий"]
            new_values = {}
            for col in editable_cols:
                new_value = input(f"Введите новое значение для колонки '{col}' (оставьте пустым, чтобы не изменять): ")
                if new_value:
                    new_values[col] = new_value

            # Обновляем строки в DataFrame
            for index in rows_to_edit.index:
                for col, new_value in new_values.items():
                    df.at[index, col] = new_value

            print("Записи отредактированы.")
            df.to_excel(file_path, sheet_name='Sheet', index=False)
            return df
        else:
            print("Записи не отредактированы.")
            return df
    else:
        print("Записи не найдены.")
        return df


def main_menu(dir_path: str, file_name: str) -> None:
    """
    Запускает меню базы абонентов

    :param dir_path: Путь к директории с файлом
    :param file_name: Имя файла
    :return: None
    """

    window = 30
    menu = {
        "0": ("Выйти из программы", sys.exit),
        "1": ("Информация о базе", info_book(dir_path, file_name)),
        "2": ("Прочитаь базу", read_book(dir_path, file_name)),
        "3": ("Добавить абонента", insert_row(dir_path, file_name)),
        "4": ("Найти абонента", lambda: print(search_rows(dir_path, file_name))),
        "5": ("Удалить абонента", lambda: remove_rows(dir_path, file_name)),
        "6": ("Изменить запись", lambda: edit_row(dir_path, file_name)),

    }

    while True:
        print("-" * window * 4)
        print(
            "***********************************Вы работаете в телефонном справочнике.***********************************\n\nМеню:")
        for key, value in menu.items():
            print(f"{key}. {value[0]}")

        print("-" * window * 4)
        choice = input("Выберите пункт меню (указав номер): ")
        if choice in menu:
            menu[choice][1]()
        else:
            print("Неверный выбор. Пожалуйста, попробуйте снова.")


# Пример использования функции
if __name__ == "__main__":
    PATH = '/db'
    FILE = 'tel_book.xlsx'
    main_menu(PATH, FILE)
